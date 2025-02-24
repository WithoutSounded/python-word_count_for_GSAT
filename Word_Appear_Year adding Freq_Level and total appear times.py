import openpyxl
import csv

translator = str.maketrans('', '', '_[]0123456789')
fieldnames = [str(year) for year in range(83,108)]
years = fieldnames
fieldnames = ['words'] + ['total'] +["Freq Level"]+ fieldnames

file = 'D:\\Justin\'s_University\\專題(learning)\\學測\\學測字 Freq Level.xlsx'
wb = openpyxl.load_workbook(file)
ws = wb.get_sheet_by_name('工作表1')

Freq_Level = dict()
for i in range(2, ws.max_row+1):
    value = ws['B'+str(i)].value
    if value is None:
        Freq_Level[ws['A'+str(i)].value] = []
    else:
        word = value.split(' ')
        word_list = []
        for j in word:
            j = j.translate(translator)
            j = j.strip('\n')
            word_list.append(j)
    Freq_Level[ws['A'+str(i)].value] = word_list
    
    with open('D:\\Justin\'s_University\\專題(learning)\\Software\\Word_Count\\output\\'+'Word_Appear_Years'+'.csv') as csvFile:
    csvReader = csv.reader(csvFile)
    csvDictReader = csv.DictReader(csvFile)
    with open('D:\\Justin\'s_University\\專題(learning)\\Software\\Word_Count\\output\\'+'Word_Appear_Years_f'+'.csv', 'w', newline='') as file:
        dictWriter = csv.DictWriter(file, fieldnames = fieldnames)
        dictWriter.writeheader()
    
        for row in csvDictReader:
            ctr = 0
            print(row['words'])
            output = row
            output['total'] = 0
            total = 0
            for year in years:
                if output[year] == '':
                    total = 0
                    output['total'] += total
                else:
                    total = 1
                    output['total'] += total
            print("the frequence total is {} ".format(output['total']))
            for freq, word in Freq_Level.items():
                if row['words'] in word:
                    output['Freq Level'] = freq
                    print("{} is Level {}".format(row['words'],freq))
                    ctr = 1
                    break
            if ctr == 0:
                output['Freq Level'] = ''
                print("{} is out of Level".format(row['words']))
            dictWriter.writerow(output)
            
            
            
